import os
import time
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Color, Font, colors
from openpyxl.styles.colors import RED, DARKYELLOW, BLUE

test = openpyxl.load_workbook('Z1_TEST.xlsx')
td = test.get_sheet_by_name('td')
hive = test.get_sheet_by_name('hive')
_h = test.get_sheet_by_name('_h')

# def changeDir():
#     zone = input("what zone? ")
#     zone = zone.upper()
#
#     dirPath = 'C:/Users/mlgkbart/Desktop/HDP/'
#     os.chdir(dirPath)
#     folder_list = os.listdir(dirPath)
#     for folders, sub_folders, file in os.walk(dirPath):
#         for name in file:
#             if name.endswith(".xlsx") and name.startswith(zone):
#                 filename = os.path.join(folders, name)
                # test = openpyxl.load_workbook(filename)
#                 ws=test.active
#                 print(name)

def prettyIter(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]

def create_dict():

    #this is retarded
    tdCompId = prettyIter(td)
    # hiveCompId = hive['B{}'].value
    # _hCompId = _h['B{}'].value
    tdCount = prettyiter(td)
    # hiveCount = hive['A{}'].value
    # _hCount = _h['A'].value

    tdThing = {tdCompId:tdCount}
    # hiveThing = {hiveCompId:hiveCount}
    # _hThing = {_hCompId:_hCount}

    print(tdThing)


# colorMap = {'hive':RED, 'td':DARKYELLOW, '_h':BLUE}
#

#
# def diffTool(sheet1, sheet2, wb): # check if hasSameValues return empty array
#     results = wb.create_sheet(sheet1.title + "_" + sheet2.title)
#
#     sheet1_list = list(prettyIter(sheet1))
#     sheet2_list = list(prettyIter(sheet2))
#     if not hasSameColumns(sheet1, sheet2):
#         results["A1"].value = "Columns is wrong"
#
#     # elif not hasSameRows(sheet1, sheet2):
#     #     results["A1"].value = "Rows wrong"
#
#     else:
#         errors = hasSameValues(sheet1_list, sheet2_list)
#         for idx in errors:
#             results.append(sheet1_list[idx])
#             results.append(sheet2_list[idx])
#
#         counter = 0
#
#         for result in results.iter_rows():
#             for cell in result:
#                 if counter %2 == 0:
#                     cell.font = Font(color=colorMap[sheet1.title])
#                 else:
#                     cell.font = Font(color=colorMap[sheet2.title])
#             counter +=1
#
#
#
# def hasSameColumns(sheet1, sheet2):
#     return sheet1.max_column == sheet2.max_column
#
# def hasSameRows(sheet1, sheet2):
#     return sheet1.max_row == sheet2.max_row
#
# def hasSameValues(sheet1_list, sheet2_list): #return array of indeces of error rows
#     output = []
#
#     for i in range(0, len(sheet1_list)):
#         for j in range(0, len(sheet1_list[i])):
#             if sheet1_list[i][j] != sheet2_list[i][j]:
#                 output.append(i)
#             else
#                 break
#
#     return output
#

if __name__ == "__main__":
    create_dict()


# results = Workbook()
#     ws = wb.active()
#     ws['A1'].value = "td results are yellow"
#
#     # diffTool(hive, td, results)
#     diffTool(td, _h, results)
#     # diffTool(hive, _h, results)
#
#
#     results.save("results_" + time.strftime("%m%d%I%M") + ".xlsx")
